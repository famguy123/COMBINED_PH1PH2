
import streamlit as st
from openpyxl import load_workbook
from copy import copy
import tempfile

def inject_sheets(template_wb, user_wb, sheet_names):
    for sheet in sheet_names:
        if sheet in user_wb.sheetnames:
            del user_wb[sheet]

    for sheet_name in sheet_names:
        source_ws = template_wb[sheet_name]
        new_ws = user_wb.create_sheet(title=sheet_name)

        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        for row_idx, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

    return user_wb

st.title("📊 Combined Sheet Injector")

st.markdown("""
Step 1: Upload the main Excel file you want to modify.

Step 2: Upload the **monthly sheets template** (contains sheets starting with `2025`).

Step 3: Upload the **summary sheets template** (contains `Γενικό Αποτέλεσμα` and `Διαφορές`).

This app will first inject the monthly sheets, and then the summary sheets, preserving formatting from both templates.
""")

user_file = st.file_uploader("📁 Upload Your Excel File", type=["xlsx"])
monthly_template = st.file_uploader("📄 Upload Monthly Template (e.g., template_phase_2_cleaned.xlsx)", type=["xlsx"])
summary_template = st.file_uploader("📄 Upload Summary Template (e.g., bilio_with_v3_formulas - Copy.xlsx)", type=["xlsx"])

if user_file and monthly_template and summary_template:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_user,          tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_output:

        tmp_user.write(user_file.read())
        tmp_user.flush()

        # Load workbooks
        user_wb = load_workbook(tmp_user.name)
        monthly_wb = load_workbook(monthly_template, data_only=False)
        summary_wb = load_workbook(summary_template, data_only=False)

        # Inject monthly sheets
        monthly_sheets = [name for name in monthly_wb.sheetnames if name.startswith("2025")]
        user_wb = inject_sheets(monthly_wb, user_wb, monthly_sheets)

        # Inject summary sheets
        summary_sheets = ["Γενικό Αποτέλεσμα", "Διαφορές"]
        user_wb = inject_sheets(summary_wb, user_wb, summary_sheets)

        # Save the final output
        user_wb.save(tmp_output.name)

        with open(tmp_output.name, "rb") as f:
            st.success("✅ Final file is ready with both monthly and summary sheets.")
            st.download_button("📥 Download Final File", f, file_name="final_combined_file.xlsx")
