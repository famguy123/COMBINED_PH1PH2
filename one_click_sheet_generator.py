import streamlit as st
from openpyxl import load_workbook
from copy import copy
import tempfile

def inject_sheets(template_wb, user_wb, sheet_names, exclude_sheets=None):
    if exclude_sheets is None:
        exclude_sheets = []

    # Delete sheets only if they are not in the exclude list
    for sheet in sheet_names:
        if sheet in user_wb.sheetnames and sheet not in exclude_sheets:
            del user_wb[sheet]

    for sheet_name in sheet_names:
        if sheet_name in exclude_sheets:
            continue

        source_ws = template_wb[sheet_name]
        new_ws = user_wb.create_sheet(title=sheet_name)

        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        for row_idx, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height
            new_ws.row_dimensions[row_idx].hidden = row_dim.hidden
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

    return user_wb

def detect_year(user_wb):
    """Detect whether the uploaded file is 2025 or 2026 based on sheet names."""
    for name in user_wb.sheetnames:
        if name.startswith("2026_"):
            return "2026"
    return "2025"

TEMPLATES = {
    "2025": {
        "monthly": "monthly_template_2025.xlsx",
        "summary": "summary_template_2025.xlsx",
    },
    "2026": {
        "monthly": "monthly_template_2026.xlsx",
        "summary": "summary_template_2026.xlsx",
    },
}

st.title("📂 One-Click Sheet Generator")

st.markdown("""
Upload your main Excel file. This app will automatically:

- Detect whether your file is **2025** or **2026**
- Inject the matching **monthly sheets** from the internal template
- Inject the **Γενικό Αποτέλεσμα** and **Διαφορές** sheets from the summary template

⚠️ The **ΕΣΟΔΑ** and **60-69 ΕΞΟΔΑ+ΟΜ 2** sheets in your uploaded file will be preserved and NOT overwritten.
""")

uploaded_file = st.file_uploader("📁 Upload Your Excel File", type=["xlsx"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_user, tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_output:

        tmp_user.write(uploaded_file.read())
        tmp_user.flush()

        user_wb = load_workbook(tmp_user.name)

        # Auto-detect year from uploaded file
        year = detect_year(user_wb)
        st.info(f"📅 Detected year: **{year}**")

        monthly_wb = load_workbook(TEMPLATES[year]["monthly"], data_only=False)
        summary_wb = load_workbook(TEMPLATES[year]["summary"], data_only=False)

        # Define sheets to exclude from deletion
        exclude_sheets = [f"{year}_ΕΣΟΔΑ", f"{year}_60-69 ΕΞΟΔΑ+ΟΜ 2"]

        # Inject monthly sheets
        monthly_sheets = [name for name in monthly_wb.sheetnames if name.startswith(year)]
        user_wb = inject_sheets(monthly_wb, user_wb, monthly_sheets, exclude_sheets=exclude_sheets)

        # Inject summary sheets
        summary_sheets = ["Γενικό Αποτέλεσμα", "Διαφορές"]
        user_wb = inject_sheets(summary_wb, user_wb, summary_sheets)

        user_wb.save(tmp_output.name)

        with open(tmp_output.name, "rb") as f:
            st.success(f"✅ File ready with all {year} sheets injected.")
            st.download_button("📥 Download Final File", f, file_name="generated_output.xlsx")
